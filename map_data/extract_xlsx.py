from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path
import json


xlsx_path = Path("voedselbos_species.xlsx")
json_path = Path("voedselbos_species.json")

print(f"Loading data from '{xlsx_path}''")
wb = load_workbook(filename=xlsx_path)
ws = wb.active

species = ws["A"]
name_la = ws["B"]
name_nl = ws["C"]
name_en = ws["D"]
height = ws["E"]
width = ws["F"]
phase = ws["G"]
notes = ws["J"]

species_list = [
    {
        "species": "Onbekend",
        "name_la": "Onbekend",
        "name_nl": "Onbekend",
        "name_en": "Unknown",
        "height": None,
        "width": None,
    }
]

for i in range(1, len(species)):

    species_list.append(
        {
            "species": species[i].value,
            "name_la": name_la[i].value,
            "name_nl": name_nl[i].value,
            "name_en": name_en[i].value,
            "height": height[i].value if height[i].value else None,
            "width": width[i].value if width[i].value else None,
        }
    )

json_data = {"species": species_list, "updated": datetime.now().isoformat()}

print(f"Saving to '{json_path}'")
with open(json_path, "w") as f:
    json.dump(json_data, f, indent=4)
